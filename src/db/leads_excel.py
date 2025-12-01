"""Функции для работы с Excel файлом leads.xlsx для сохранения данных о лидах"""
import logging
import os
import asyncio
import time
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

from ..config import settings
from .models import UserProfile

logger = logging.getLogger(__name__)


def _get_excel_file_path() -> str:
    """
    Получить путь к Excel файлу leads.xlsx.
    
    Returns:
        Путь к файлу (по умолчанию в корне проекта)
    """
    excel_path = settings.leads_excel_path
    if not excel_path:
        # По умолчанию в корне проекта
        project_root = Path(__file__).resolve().parent.parent.parent
        excel_path = str(project_root / "leads.xlsx")
    else:
        # Если указан .xls, заменяем на .xlsx
        if excel_path.endswith('.xls') and not excel_path.endswith('.xlsx'):
            excel_path = excel_path[:-4] + '.xlsx'
    return excel_path


def _sync_save_to_excel(profile: UserProfile, name_sys: str = "") -> Optional[str]:
    """
    Синхронная функция для сохранения данных лида в Excel файл.
    
    Args:
        profile: Профиль пользователя из UserProfile
        name_sys: Системное имя пользователя (first_name или username)
        
    Returns:
        Путь к сохраненному файлу или None в случае ошибки
    """
    try:
        try:
            from openpyxl import Workbook, load_workbook  # type: ignore
            from openpyxl.styles import Font, Alignment, Border, Side  # type: ignore
        except ImportError as import_err:
            logger.error(f"❌ Не удалось импортировать openpyxl: {import_err}", exc_info=True)
            raise
        
        excel_path = _get_excel_file_path()
        logger.info(f"📁 Путь к Excel файлу: {excel_path}")
        
        # Создаем директорию, если её нет
        excel_dir = os.path.dirname(excel_path)
        if excel_dir and not os.path.exists(excel_dir):
            os.makedirs(excel_dir, exist_ok=True)
            logger.info(f"📁 Создана директория: {excel_dir}")
        
        # Проверяем, существует ли файл
        file_exists = os.path.exists(excel_path)
        logger.info(f"📄 Файл существует: {file_exists}")
        headers = ["Дата Date", "Статус Status", "Имя Name", "Систем. Имя Name_sys", "Телефон Phone", "Опыт Exp", "Уровень Level", "Цели Goals", "Ранее Before", "Политика Politic"]
        
        if file_exists:
            try:
                # Загружаем файл без keep_vba, чтобы избежать проблем с повреждением
                # Используем data_only=False для сохранения формул (если есть)
                workbook = load_workbook(excel_path, read_only=False, data_only=False)
                worksheet = workbook.active
                # Проверяем, есть ли заголовки в первой строке
                first_row = [str(cell.value).strip() if cell.value else "" for cell in worksheet[1]]
                # Проверяем, совпадают ли заголовки (с учетом возможных русских заголовков)
                # Если заголовки не совпадают или файл содержит только заголовки, обновляем их
                if not first_row or first_row != headers:
                    # Если в файле только заголовки (1 строка), просто заменяем их
                    if worksheet.max_row == 1:
                        # Очищаем первую строку и добавляем правильные заголовки
                        for col_idx in range(1, len(headers) + 1):
                            cell = worksheet.cell(row=1, column=col_idx)
                            cell.value = headers[col_idx - 1] if col_idx <= len(headers) else None
                        # Удаляем лишние ячейки, если они есть
                        if worksheet.max_column > len(headers):
                            worksheet.delete_cols(len(headers) + 1, worksheet.max_column - len(headers))
                    else:
                        # Если есть данные, заменяем только заголовки
                        worksheet.delete_rows(1)
                        worksheet.insert_rows(1)
                        for col_idx, header in enumerate(headers, start=1):
                            cell = worksheet.cell(row=1, column=col_idx)
                            cell.value = header
                    logger.info(f"Заголовки в файле {excel_path} обновлены на: {headers}")
            except Exception as e:
                logger.warning(f"Не удалось открыть существующий файл {excel_path}, создаем новый: {e}")
                workbook = Workbook()
                worksheet = workbook.active
                # Добавляем заголовки
                worksheet.append(headers)
        else:
            # Создаем новый файл
            workbook = Workbook()
            worksheet = workbook.active
            # Добавляем заголовки
            worksheet.append(headers)
        
        # Форматируем дату
        date_str = profile.date.strftime("%Y-%m-%d %H:%M:%S") if profile.date else datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        
        # Используем name_sys из профиля, если переданный пустой
        final_name_sys = name_sys or profile.name_sys or ""
        
        # Если Name не указан, используем Name_sys
        final_name = profile.name or final_name_sys or ""
        
        # Собираем данные (все значения, даже пустые, для соответствия структуре)
        row_data = [
            date_str,
            profile.status or "",
            final_name,
            final_name_sys,
            profile.phone or "",
            profile.exp or "",
            profile.level or "",
            profile.goals or "",
            profile.before or "",
            profile.politic or "",
        ]
        
        # Проверяем на дубликаты по всем строкам файла
        # Дубликат = та же комбинация статуса, имени, системного имени и телефона
        is_duplicate = False
        if worksheet.max_row > 1:  # Если есть данные кроме заголовков
            # Проверяем все строки начиная со строки 2 (строка 1 - заголовки)
            for row_idx in range(2, worksheet.max_row + 1):
                try:
                    existing_status = str(worksheet.cell(row=row_idx, column=2).value or "").strip()
                    existing_name = str(worksheet.cell(row=row_idx, column=3).value or "").strip()
                    existing_name_sys = str(worksheet.cell(row=row_idx, column=4).value or "").strip()
                    existing_phone = str(worksheet.cell(row=row_idx, column=5).value or "").strip()
                    
                    new_status = str(row_data[1]).strip()
                    new_name = str(row_data[2]).strip()
                    new_name_sys = str(row_data[3]).strip()
                    new_phone = str(row_data[4]).strip()
                    
                    # Считаем дубликатом, если статус, имя, системное имя и телефон совпадают
                    if (existing_status == new_status and 
                        existing_name == new_name and 
                        existing_name_sys == new_name_sys and
                        existing_phone == new_phone and
                        new_status and new_name and new_name_sys and new_phone):  # Только если все четыре поля заполнены
                        is_duplicate = True
                        logger.info(
                            f"⚠️ Обнаружен дубликат в строке {row_idx}: "
                            f"статус='{new_status}', имя='{new_name}', системное имя='{new_name_sys}', телефон='{new_phone}'. Пропускаем добавление."
                        )
                        break
                except Exception as e:
                    logger.warning(f"Ошибка при проверке дубликата в строке {row_idx}: {e}")
                    continue
        
        if is_duplicate:
            logger.info(f"⏭️ Дубликат не добавлен в Excel для пользователя {profile.tg_user_id}")
            return None  # Не добавляем дубликат, не отправляем email
        
        # Добавляем строку
        logger.debug(f"Добавление строки в Excel: {row_data}")
        worksheet.append(row_data)
        
        # Применяем форматирование ко всем ячейкам
        # Создаем стили
        font = Font(size=12)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Применяем форматирование ко всем ячейкам в файле
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.font = font
                cell.alignment = alignment
                cell.border = thin_border
        
        # Сохраняем файл с повторными попытками
        max_retries = 5
        retry_delay = 0.5  # секунды
        
        saved_successfully = False
        temp_path = excel_path + ".tmp"
        
        # Сначала пытаемся сохранить напрямую в основной файл
        for attempt in range(1, max_retries + 1):
            try:
                workbook.save(excel_path)
                saved_successfully = True
                break
            except PermissionError:
                if attempt < max_retries:
                    wait_time = retry_delay * attempt
                    logger.warning(
                        f"⚠️ Файл {excel_path} заблокирован (попытка {attempt}/{max_retries}). "
                        f"Повторная попытка через {wait_time} сек. Закройте файл в Excel, если он открыт."
                    )
                    time.sleep(wait_time)
                else:
                    # Если все попытки не удались, сохраняем во временный файл
                    logger.warning(
                        f"⚠️ Не удалось сохранить в основной файл {excel_path} после {max_retries} попыток. "
                        f"Сохраняю во временный файл {temp_path}."
                    )
                    try:
                        workbook.save(temp_path)
                        logger.warning(
                            f"⚠️ Данные сохранены во временный файл {temp_path}. "
                            f"Закройте {excel_path} в Excel и переименуйте {temp_path} в {excel_path} вручную, "
                            f"или удалите {excel_path} и переименуйте {temp_path}."
                        )
                        # Не поднимаем исключение - данные сохранены, просто не в основной файл
                        saved_successfully = True
                    except Exception as temp_error:
                        logger.error(
                            f"❌ Не удалось сохранить даже во временный файл {temp_path}: {temp_error}",
                            exc_info=True
                        )
                        raise
            except Exception as save_error:
                logger.error(f"❌ Ошибка при сохранении файла {excel_path}: {save_error}", exc_info=True)
                raise
        
        if saved_successfully:
            final_path = excel_path if os.path.exists(excel_path) else temp_path
            logger.info(
                f"✅ Данные лида сохранены в Excel файл {final_path} для пользователя {profile.tg_user_id}, "
                f"статус: {profile.status}, имя: {profile.name or 'не указано'}, "
                f"телефон: {profile.phone or 'не указан'}, строк в файле: {worksheet.max_row}"
            )
            return final_path
        return None
        
    except ImportError as import_err:
        logger.error(f"❌ Библиотека openpyxl не установлена или не может быть импортирована: {import_err}", exc_info=True)
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка при сохранении данных в Excel файл: {e}", exc_info=True)
        raise


async def save_lead_to_excel(profile: UserProfile, name_sys: str = "") -> None:
    """
    Асинхронная функция для сохранения данных лида в Excel файл.
    После успешного сохранения отправляет файл на email из EMAIL_MAIN.
    
    Args:
        profile: Профиль пользователя из UserProfile
        name_sys: Системное имя пользователя (first_name или username)
    """
    try:
        logger.info(f"🔄 Начало сохранения в Excel для пользователя {profile.tg_user_id}, статус: {profile.status}")
        # Запускаем синхронную функцию в отдельном потоке
        loop = asyncio.get_event_loop()
        saved_file_path = await loop.run_in_executor(None, _sync_save_to_excel, profile, name_sys)
        logger.info(f"✅ Успешно завершено сохранение в Excel для пользователя {profile.tg_user_id}")
        
        # Отправляем файл на email после успешного сохранения
        if saved_file_path:
            try:
                from ..email_sender import send_email_with_attachment
                await send_email_with_attachment(
                    file_path=saved_file_path,
                    subject=f"Обновление leads.xlsx - новый лид",
                    body=f"Файл leads.xlsx был обновлен.\n\nДанные лида:\n- Статус: {profile.status or 'не указан'}\n- Имя: {profile.name or profile.name_sys or 'не указано'}\n- Телефон: {profile.phone or 'не указан'}\n\nСм. вложение."
                )
            except Exception as email_error:
                # Не прерываем выполнение, если отправка email не удалась
                logger.error(f"❌ Ошибка при отправке email: {email_error}", exc_info=True)
    except Exception as e:
        logger.error(f"❌ Ошибка при запуске сохранения в Excel для пользователя {profile.tg_user_id}: {e}", exc_info=True)
        raise
